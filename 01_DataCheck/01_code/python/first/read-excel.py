# -*- coding: utf-8 -*-

import pymysql,time

import datetime

from openpyxl import load_workbook

file_name = input("请输入文件名：") # 不包括文件扩展名，与Python脚本在同一目录

time_load = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

print('>> 当前时间：',time_load)

print (" + 正在处理：" + file_name)

time_start = time.time() # 记录初始时间，用于计算单个文件的处理时长

wb = load_workbook(filename = file_name+'.xlsx') # 打开文件

ws = wb[wb.sheetnames[0]] # 选择sheet表

rows = ws.max_row # 获取表的最大行数

columns = ws.max_column # 获取表的最大列数

column_heading = [ws.cell(row=1,column=x).value for x in range(1,columns+1)] # 读取excel第一行的值，写入list

column_name = ['URL', '城市', '项目名称', '坐落位置', '开发企业', '预售许可证编号', '发证日期', '开盘日期', '预售证准许销售面积', '销售状态', '销售楼号', '套数', '面积', '拟售价格', '售楼电话', '售楼地址', '房号', '房屋建筑面积', '房屋销售状态（颜色区分）'
] # 数据库必需字段

if len([name for name in column_name if name not in column_heading]) == 0: # 返回字段组成的list为空，则说明文件列标题包含MySQL需要的字段

 print(' - 检查完成，执行写入')

# 判断Excel中各字段所在列号
url = column_heading.index(column_name[0])
city_name = column_heading.index(column_name[1])
floor_name = column_heading.index(column_name[2])
address = column_heading.index(column_name[3])
business = column_heading.index(column_name[4])
issue_code = column_heading.index(column_name[5])
issue_date = column_heading.index(column_name[6])
open_date = column_heading.index(column_name[7])
issue_area = column_heading.index(column_name[8])
sale_state = column_heading.index(column_name[9])
building_code = column_heading.index(column_name[10])
room_sum = column_heading.index(column_name[11])
area = column_heading.index(column_name[12])
simulation_price = column_heading.index(column_name[13])
sale_telephone = column_heading.index(column_name[14])
sale_address = column_heading.index(column_name[15])
room_code = column_heading.index(column_name[16])
room_sale_area = column_heading.index(column_name[17])
room_sale_state = column_heading.index(column_name[18])

data = []

for row in range(2, rows + 1):


 data.append(str(ws.cell(row=row,column=1).value))
 print(*data,sep= '\n')


print(' - 完成写入')