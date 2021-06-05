# -*- coding: utf-8 -*-

import pymysql,time

from openpyxl import load_workbook

# 1.0准备

# 1.1指定xlsx文件

file_name = input("请输入文件名：") # 不包括文件扩展名，与Python脚本在同一目录

# 1.2记录提交时间

time_load = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

print('>> 当前时间：',time_load)

# print('\n',end='')

# 1.3指定需要存入MySQL的产品列表(三级分类)

#List_Product = ['手机','PC','配件']

# 2.0链接数据库(注意把账户和密码改成自己的)

print('>> 连接MySQL……')

conn = pymysql.connect(

host='172.28.36.77',

port=3306,

user='mysql',

passwd='egSQ7HhxajHZjvdX',

db='temp_db',

charset='utf8')

def connect_mysql(conn):

#判断链接是否正常

 conn.ping(True)

#建立操作游标

 cursor=conn.cursor()

#设置数据输入输出编码格式

 cursor.execute('set names utf8')

 return cursor

# 建立链接游标

cur=connect_mysql(conn)

# 写入数据库时使用的语句

insert_sql = 'insert into tmp_city_newest_deal (\
    url,city_name,floor_name,address,business,issue_code,issue_date,open_date,issue_area,sale_state,building_code,room_sum,area,simulation_price,sale_telephone,sale_address,room_code,room_sale_area,room_sale_state\
        ) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s , %s, %s, %s)'

# 3.0摘取文件的数据提交mysql

print (" + 正在处理：" + file_name)

time_start = time.time() # 记录初始时间，用于计算单个文件的处理时长

wb = load_workbook(filename = file_name+'.xlsx') # 打开文件

ws = wb[wb.sheetnames[0]] # 选择sheet表

rows = ws.max_row # 获取表的最大行数

columns = ws.max_column # 获取表的最大列数

column_heading = [ws.cell(row=1,column=x).value for x in range(1,columns+1)] # 读取excel第一行的值，写入list

column_name = ['URL', '城市', '项目名称', '坐落位置', '开发企业', '预售许可证编号', '发证日期', '开盘日期', '预售证准许销售面积', '销售状态', '销售楼号', '套数', '面积', '拟售价格', '售楼电话', '售楼地址', '房号', '房屋建筑面积', '房屋销售状态（颜色区分）'
] # 数据库必需字段

if len([name for name in column_name if name not in column_heading]) == 0: # 返回字段组成的list为空，则说明文件列标题包含MySQL需要的字段

 print(' - 检查完成，执行写入')

# 判断Excel中各字段所在列号
url = column_heading.index(column_name[0])
city_name = column_heading.index(column_name[1])
floor_name = column_heading.index(column_name[2])
address = column_heading.index(column_name[3])
business = column_heading.index(column_name[4])
issue_code = column_heading.index(column_name[5])
issue_date = column_heading.index(column_name[6])
open_date = column_heading.index(column_name[7])
issue_area = column_heading.index(column_name[8])
sale_state = column_heading.index(column_name[9])
building_code = column_heading.index(column_name[10])
room_sum = column_heading.index(column_name[11])
area = column_heading.index(column_name[12])
simulation_price = column_heading.index(column_name[13])
sale_telephone = column_heading.index(column_name[14])
sale_address = column_heading.index(column_name[15])
room_code = column_heading.index(column_name[16])
room_sale_area = column_heading.index(column_name[17])
room_sale_state = column_heading.index(column_name[18])

#kssj = column_heading.index(column_name[0]) # 开始时间 - 位置

#yhdh = column_heading.index(column_name[1]) # 来访ID - 位置

#fl2j = column_heading.index(column_name[2]) # 二级分类 - 位置

#fl3j = column_heading.index(column_name[3]) # 三级分类 - 位置

#fl4j = column_heading.index(column_name[4]) # 四级分类 - 位置

#fl5j = column_heading.index(column_name[5]) # 五级分类 - 位置

#kfpg = column_heading.index(column_name[6]) # 备注 - 位置

# 判定第2行是否为合并行，以此确定选择的表格区域(如果出现空行，则会导致自动选区范围不是表格内容覆盖的最大值区域)。

if ws.cell(row=2,column=1).value == None:

 table_start_line = 3

else:

 table_start_line = 2

 data = []

for row in range(table_start_line, rows + 1):

# 此处用于筛选“三级分类”的关键字

#if ws.cell(row=row, column=fl3j+1).value not in List_Product: # 判断是否为所需行

#pass

#else:

 for column in range(1, columns+1): # 因为从第1列开始，所以此处从1开始

  data.append(str(ws.cell(row=row, column=column).value)) # 以字符串形式保存数据到MySQL

 cur.execute(insert_sql, (data[url], data[city_name], data[floor_name], data[address], data[business], data[issue_code], data[issue_date], data[open_date], data[issue_area], data[sale_state], data[building_code], data[room_sum], data[area], data[simulation_price], data[sale_telephone], data[sale_address], data[room_code], data[room_sale_area], data[room_sale_state]))

 data = [] # 每写一行则重置为空

print(' - 完成写入')

#else:

#print(' - 文件列标题不完全包含MySQL需要的字段，请检查文件。')

wb.close() # 关闭excel

time_use = time.time() - time_start # 计算处理文件的时长

print(' - 用时：' + str(time_use) + '秒') # 单位默认是秒

# 4.0结束

conn.commit() # 提交记录

conn.close() # 关闭数据库链接

print('>> Done!') #完毕